import os
from mysql.connector import Error
from conexionBD import conectar
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import re

def borrar_pantalla():
    os.system('cls' if os.name == 'nt' else 'clear')

def esperar_tecla():
    input("\nPresione ENTER para continuar...")

def menu_principal():
    borrar_pantalla()
    print("\n" + "=" * 50)
    print(" SISTEMA GESTIÓN COOPPEL ")
    print("=" * 50)
    print("1. Iniciar sesión")
    print("2. Registrar usuario")
    print("3. Salir")
    return input("\nSeleccione una opción: ")

def menu_inventario():
    borrar_pantalla()
    print("\n" + "=" * 50)
    print(" MENÚ PRINCIPAL - COOPPEL")
    print("=" * 50)
    print("1. Realizar venta")
    print("2. Registrar producto")
    print("3. Listar productos")
    print("4. Buscar producto")
    print("5. Actualizar stock")
    print("6. Menú de administración de tablas")
    print("7. Cerrar sesión")
    return input("\nSeleccione una opción: ")

def menu_admin():
    borrar_pantalla()
    print("\n" + "=" * 50)
    print(" MENÚ DE ADMINISTRACIÓN DE TABLAS")
    print("=" * 50)
    print("1. Reiniciar tabla de productos")
    print("2. Reiniciar tabla de ventas")
    print("3. Reiniciar tabla de detalle_venta")
    print("4. Reiniciar tabla de usuarios")
    print("5. Exportar tablas a PDF y/o Excel...")
    print("6. Volver al menú principal")
    return input("\nSeleccione una opción: ")

def formatear_precio(precio):
    return f"${precio:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def tabla_excel(datos, nombre_archivo="reporte"):
    if not datos:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte COOPPEL"
    encabezados = list(datos[0].keys())
    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col, value=encabezado)
        cell.font = Font(bold=True)
    for r, item in enumerate(datos, start=2):
        for c, key in enumerate(encabezados, start=1):
            ws.cell(row=r, column=c, value=item.get(key, ""))
    for column in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = max(10, min(40, max_len+2))
    os.makedirs("reportes", exist_ok=True)
    nombre_completo = f"{nombre_archivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta = os.path.join("reportes", nombre_completo)
    wb.save(ruta)
    return ruta

def tabla_pdf(datos, nombre_archivo="reporte"):
    if not datos:
        return None
    os.makedirs("reportes", exist_ok=True)
    nombre_archivo = re.sub(r'[^\w-]', '', nombre_archivo)
    ruta = os.path.join("reportes", f"{nombre_archivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

    doc = SimpleDocTemplate(ruta, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Reporte COOPPEL", styles['Title']))
    elements.append(Spacer(1, 12))

    encabezados = [str(k) for k in datos[0].keys()]
    filas = []
    for item in datos:
        fila = [str(item.get(k, ""))[:80] for k in encabezados]
        filas.append(fila)

    tabla_datos = [encabezados] + filas
    tabla = Table(tabla_datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F2F2F2')),
        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ]))
    elements.append(tabla)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Italic']))
    doc.build(elements)
    return ruta

def exportar_tabla(tabla, formato):
    datos = obtener_datos_tabla(tabla)
    if not datos:
        return None
    nombre_archivo = f"reporte_{tabla}"
    if formato == 'excel':
        return tabla_excel(datos, nombre_archivo)
    elif formato == 'pdf':
        return tabla_pdf(datos, nombre_archivo)
    return None

def obtener_datos_tabla(tabla):
    conn = conectar()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(f"SELECT * FROM {tabla}")
        return cur.fetchall()
    except Error as e:
        print(f"Error al obtener datos de {tabla}: {e}")
        return None
    finally:
        if conn.is_connected():
            cur.close()
            conn.close()
